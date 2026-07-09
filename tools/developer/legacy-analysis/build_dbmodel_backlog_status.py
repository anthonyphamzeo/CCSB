from __future__ import annotations

import csv
import html
import json
import re
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[3]
OUT_DIR = ROOT / "docs" / "tracking" / "status" / "dbmodel-review"
DBMODEL = ROOT / "docs" / "design" / "data-model" / "source" / "dbmodel.xlsx"
TASKS_CSV = ROOT / "tasks" / "backlog" / "import" / "CCSB_Backlog_Import_Tasks_v1_1.csv"
STORIES_CSV = ROOT / "tasks" / "backlog" / "import" / "CCSB_Backlog_Import_User_Stories_v1_1.csv"
ACTION_PLAN = ROOT / "docs" / "tracking" / "decisions" / "CCSB_Next_Action_Plan_After_Validation.html"


COMPLETE_RULES = {
    "US-013-T01": (
        "Runtime projection schema is present: ccsb_runtimeprojection has payload JSON, projection schema version, source change token, content hash, generated timestamp, status, scope and sizing fields.",
        "Database model evidence satisfies the schema-definition part of this task. Generator, validation and import hooks remain separate tasks.",
    ),
    "US-020-T01": (
        "Group schema is present: ccsb_groupdefinition, ccsb_groupnode and ccsb_groupmembership model ordered levels, materialized nodes, resolved paths, resource membership and active/effective state.",
        "Database model evidence satisfies the group-definition schema task.",
    ),
    "US-021-T03": (
        "Path identity is represented by group node/member keys plus ccsb_resolvedpath and resource membership links.",
        "Database model evidence satisfies the persistence/cache-shape portion; resolver determinism tests remain open.",
    ),
    "US-044-T01": (
        "Status lifecycle schema is present: ccsb_statusmodel, ccsb_statusdefinition and ccsb_statustransition include status keys, labels, categories, permissions/reasons and validation rule links.",
        "Database model evidence satisfies the data-model task. Command-service enforcement remains open.",
    ),
    "US-050-T01": (
        "Completion/allocation measures are present across ccsb_schedulerequirement and ccsb_scheduleassignment: required/min/max units, assigned units, role keys, priority and requirement status.",
        "Database model evidence satisfies the data-shape task. Runtime calculation and tests remain open.",
    ),
    "US-070-T02": (
        "Permission schema is present: ccsb_permissionprofile and ccsb_permissionassignment include profiles, action flags, principal type/id, board/location/group scopes and access effect.",
        "Database model evidence satisfies the profile/assignment schema task. Role packaging and negative authorization tests remain open.",
    ),
}


PROGRESS_RULES = {
    "US-004-T01": (
        "Physical schema exists with 39 custom ccsb_* tables, field types, requiredness, choices, lookup fields, ownership and form-location metadata.",
        "Do not mark complete yet: V1 gaps remain for publish lock, snapshot item, schedule change, outbox/audit, activity/capability decisions, audit enablement and approval baseline.",
    ),
    "US-004-T02": (
        "Product-owned configuration tables exist for board/version, entity-field-relationship mapping, groups, status, roles, rules, permissions, runtime projection and validation results.",
        "Do not mark complete yet: baseline forms/views are not release-ready and some solution/export/import evidence is still missing.",
    ),
    "US-004-T03": (
        "Control/publish foundations exist: ccsb_operation, ccsb_operationitem, ccsb_conflict and ccsb_publishsnapshot.",
        "Missing required V1 control tables or equivalents: publish lock, publish snapshot item, schedule change ledger, audit log/operation audit and outbox event.",
    ),
    "US-004-T04": (
        "Native schedule graph foundations exist: schedule item, assignment, requirement, resource, role, service, location, calendar, availability, leave, shift and holiday tables.",
        "Still incomplete: no separate ccsb_scheduleactivity or ccsb_resourcecapability table/equivalent approval is recorded.",
    ),
    "US-004-T05": (
        "Lifecycle and lookup fields are modeled across board, board version, schedule version, publish snapshot, status and relationship tables.",
        "Still needs enforceable alternate keys, uniqueness/cascade policy, retention and automated schema tests.",
    ),
    "US-004-T06": (
        "Entity ownership choices and permission profile/assignment tables are modeled.",
        "Still needs a signed security matrix, field-security design and negative role tests. Entity-level auditing is disabled across the workbook.",
    ),
    "US-004-T07": (
        "Query-supporting fields exist for board/version, source entity/record, resource, date ranges, publish state, scope, status and hashes.",
        "Still needs an approved query/indexing plan and performance validation.",
    ),
    "US-004-T08": (
        "Validation support exists through ccsb_configurationvalidationresult, validation status fields and operation/operation item evidence tables.",
        "Still needs live metadata validation implementation and compatibility checks.",
    ),
    "US-002-T01": (
        "Versioning concepts are modeled on board version and runtime projection.",
        "Still missing explicit semantic product/schema compatibility contract and supported upgrade-path policy.",
    ),
    "US-002-T02": (
        "Board version has version number/label/configuration hash; runtime projection has projection schema version, source change token and content hash.",
        "Still missing explicit product version, schema version and migration-state fields or approved equivalents.",
    ),
    "US-002-T03": (
        "Pre-activation validation has a persistence target via ccsb_configurationvalidationresult and board-version validation fields.",
        "Still needs executable fail-closed validation logic and diagnostic code mapping.",
    ),
    "US-002-T04": (
        "Validation result and runtime projection status/error fields provide a data basis for diagnostics.",
        "Still needs administrator-facing diagnostics UX and safe re-validation commands.",
    ),
    "US-003-T01": (
        "Board registry fields exist: board code, active board version, lifecycle/status, time zone, runtime enablement and effective dates.",
        "Still missing explicit environment-boundary fields/keys and proof of uniqueness per Dataverse environment.",
    ),
    "US-003-T04": (
        "Board active version and board-version lifecycle fields are present.",
        "Still needs activation/retirement validation that prevents zero/multiple active versions.",
    ),
    "US-010-T01": (
        "Board-version aggregate fields exist: lifecycle, validation status, immutable flag, clone-from/supersedes, activation, effective dates and configuration hash.",
        "Still needs enforcement of single active version and immutability beyond modeled fields.",
    ),
    "US-010-T03": (
        "Board-version fields needed by administration forms/views exist.",
        "Still depends on the Model-Driven UX pilot because current bulk UX generation is not release-ready.",
    ),
    "US-010-T04": (
        "Immutability and lifecycle fields exist on board version.",
        "Still needs server-side transition enforcement and audit evidence.",
    ),
    "US-011-T01": (
        "Entity, field and relationship mapping tables exist with semantic role/name, source field/type, transformation, lookup target and validation metadata.",
        "Still needs approved semantic catalogue records for Booking, Activity, Assignment and Resource.",
    ),
    "US-011-T04": (
        "Mapping tables include metadata ETag, validation status, required/read/write fields and lookup target metadata.",
        "Still needs executable graph cardinality/type validation.",
    ),
    "US-011-T05": (
        "Runtime projection and normalized mapping tables provide the persistence shape for a typed resolver.",
        "Still needs resolver implementation and typed contracts.",
    ),
    "US-012-T01": (
        "Field mappings include read-enabled and write-behavior controls.",
        "Not complete: independent rollback-managed and sensitive/display-only flags are not present.",
    ),
    "US-012-T03": (
        "Field mappings include write behavior, validation expression, target data type and transformation metadata.",
        "Still needs rollback-safety rules and explicit blocked field classes.",
    ),
    "US-013-T02": (
        "Runtime projection table includes generated/status/hash/source-token fields needed by a generator.",
        "Still needs the generator implementation and idempotency behavior.",
    ),
    "US-013-T03": (
        "Projection hash, schema version, current flag, source change token and status fields support stale/incompatible detection.",
        "Still needs fail-closed runtime load checks.",
    ),
    "US-013-T04": (
        "Projection health fields exist: generated date, source token/hash, status and error detail.",
        "Still needs administration UX.",
    ),
    "US-020-T04": (
        "Group definition includes source type, ordered level number, parent definition, path separator and active state.",
        "Still needs resolver safety validation and hierarchy-limit enforcement.",
    ),
    "US-020-T05": (
        "Group node and membership tables model resolved paths and materialized hierarchy nodes.",
        "Still needs query/resolver contracts.",
    ),
    "US-021-T01": (
        "Materialized group node and membership structures provide the data basis for stable path resolution.",
        "Still needs the deterministic resolution algorithm.",
    ),
    "US-021-T02": (
        "Group membership/resource/path structure supports ambiguity detection.",
        "Still needs validation rules and diagnostics.",
    ),
    "US-022-T01": (
        "Resource role, schedule requirement, schedule assignment and resource tables model role keys, quantities, capacity and eligibility criteria.",
        "Still needs explicit approval of capability strategy because no ccsb_resourcecapability table exists.",
    ),
    "US-022-T05": (
        "Role and requirement fields support capacity/eligibility policy inputs.",
        "Still needs enforcement logic.",
    ),
    "US-030-T01": (
        "Board, active board version, runtime enabled, timezone and permission-scope fields exist.",
        "Still needs runtime bootstrap service and permission filtering.",
    ),
    "US-030-T03": (
        "Board/resource/location/calendar timezone fields exist.",
        "Still needs timezone-safe navigation implementation and DST tests.",
    ),
    "US-030-T05": (
        "Board version validation status and runtime projection status fields provide fail-closed inputs.",
        "Still needs runtime load blocker implementation.",
    ),
    "US-031-T01": (
        "Runtime projection plus schedule item/assignment/resource/date/status fields provide read-model query shape.",
        "Still needs typed query contracts and server paging.",
    ),
    "US-033-T01": (
        "Overlay source tables exist for availability windows, leave, shifts, calendars, calendar exceptions/rules and public holidays.",
        "Still needs approved overlay mapping semantics and precedence.",
    ),
    "US-033-T04": (
        "Overlay tables have active/effective/time fields for validation.",
        "Still needs optional mapping/precedence validation logic.",
    ),
    "US-040-T04": (
        "Field mappings and schedule item defaults/managed fields provide data inputs for create/edit.",
        "Still needs command service mapping and policy enforcement.",
    ),
    "US-040-T05": (
        "Schedule item contains requested/scheduled start/end, parent item, item type and scheduling status.",
        "Still needs interval and parent-rule validation.",
    ),
    "US-041-T04": (
        "Schedule assignment links resource, role, schedule item, requirement and schedule version.",
        "Still needs resolver implementation against active mappings.",
    ),
    "US-041-T05": (
        "Resource role, requirement and assignment fields provide eligibility/capacity inputs.",
        "Still needs calculation and validation logic.",
    ),
    "US-042-T04": (
        "Schedule graph and conflict tables can represent affected schedule items/assignments/resources.",
        "Still blocked by missing publish lock/change-ledger model for lock-aware mutation.",
    ),
    "US-042-T06": (
        "Operation and operation item tables provide operation evidence fields.",
        "Still needs writes from the command runtime.",
    ),
    "US-043-T01": (
        "Schedule assignment includes start and end override fields plus role/resource requirement links.",
        "Still needs board-level policy semantics for when overrides are permitted.",
    ),
    "US-043-T04": (
        "Assignment override fields exist for boundary validation.",
        "Still needs validation logic and policy flags.",
    ),
    "US-044-T02": (
        "Status model and transition records support transition policy, reason and permission requirements.",
        "Still needs functional policy definition and approval.",
    ),
    "US-044-T04": (
        "Status transition records link validation rules and required access/reason fields.",
        "Still needs command-service enforcement.",
    ),
    "US-050-T04": (
        "Requirement and assignment measures provide inputs for coverage rules.",
        "Still needs rule execution.",
    ),
    "US-051-T01": (
        "Rule definition and conflict tables model rule metadata, severity/blocking, conflict type/severity and detail JSON.",
        "Still lacks a complete confirmation-token/rule-outcome contract.",
    ),
    "US-051-T06": (
        "Conflict detail JSON and operation item error/result fields can store rule diagnostics.",
        "Still needs runtime diagnostics writes and minimization rules.",
    ),
    "US-052-T01": (
        "Persistent ccsb_conflict table exists with conflict key/type, severity, publish blocking, resource/item/assignment links and resolution fields.",
        "Still missing direct operation/correlation linkage required by supportability.",
    ),
    "US-053-T01": (
        "Schedule item/assignment statuses, publish state and requirement links provide data for unscheduled/partial criteria.",
        "Still needs query criteria and workbench semantics.",
    ),
    "US-060-T01": (
        "Publish snapshot has scope type/from/to, board/version, schedule version and manifest/state hash fields.",
        "Still needs publish-scope resolution policy and preview logic.",
    ),
    "US-060-T04": (
        "Operation and validation result tables support revalidation evidence.",
        "Still needs immediate pre-commit validation implementation.",
    ),
    "US-061-T01": (
        "Publish snapshot header exists with publish sequence, predecessor, manifest, scope, actor, operation and retention fields.",
        "Not complete: snapshot item data and managed before/after field images are missing.",
    ),
    "US-061-T03": (
        "Permission profile includes canpublish/canrollback/canoverrideconflicts and canviewaudit flags.",
        "Still needs publish-lock permission enforcement and negative tests.",
    ),
    "US-061-T06": (
        "Publish snapshot includes retention date and operation linkage.",
        "Still needs retention/access policy and operational runbook.",
    ),
    "US-062-T01": (
        "Rollback appears in operation types and publish snapshot predecessor/scope fields.",
        "Still needs compatible rollback policy and snapshot item/rollback-managed field model.",
    ),
    "US-062-T04": (
        "Operation/conflict/publish structures provide inputs for unsafe-restoration checks.",
        "Still needs lock policy, compatibility and privilege validation.",
    ),
    "US-063-T01": (
        "Schedule item/assignment publish state and publish snapshot links provide some official-state inputs.",
        "Not complete: publish lock and schedule change/pending-change model are missing.",
    ),
    "US-070-T01": (
        "Permission profiles model CCSB actions and scopes.",
        "Still needs approved role/security matrix, Dataverse role mapping and negative cases.",
    ),
    "US-071-T01": (
        "Operation/operation item tables model operation types, status, correlation, idempotency key, payloads, result/errors, counts and retry metadata.",
        "Still needs API envelope documentation and generated typed DTOs.",
    ),
    "US-071-T03": (
        "Operation includes idempotency key, attempt counts and retry metadata; Dataverse row versions are present as standard versionnumber fields.",
        "Still needs deterministic duplicate/stale-response implementation.",
    ),
    "US-072-T01": (
        "Operation and operation item tables carry correlation IDs and link to board/version/validation/publish snapshot evidence.",
        "Not complete: conflict, snapshot/change/lock/audit/outbox correlation coverage is incomplete.",
    ),
    "US-072-T02": (
        "Operation and operation item tables can persist structured operation outcome and item-level errors/results.",
        "Still needs command-transaction writes and links to audit/conflict/snapshot evidence.",
    ),
    "US-072-T04": (
        "Operation duration, result/error, item counts and scope fields provide telemetry inputs.",
        "Still needs privacy-safe telemetry schema and log review.",
    ),
}


EXPECTED_V1_TABLES = {
    "ccsb_board",
    "ccsb_boardversion",
    "ccsb_entitydefinition",
    "ccsb_fieldmapping",
    "ccsb_relationshipmapping",
    "ccsb_groupdefinition",
    "ccsb_groupnode",
    "ccsb_groupmembership",
    "ccsb_statusmodel",
    "ccsb_statusdefinition",
    "ccsb_statustransition",
    "ccsb_resourcerole",
    "ccsb_ruledefinition",
    "ccsb_permissionprofile",
    "ccsb_permissionassignment",
    "ccsb_runtimeprojection",
    "ccsb_configurationvalidationresult",
    "ccsb_operation",
    "ccsb_operationitem",
    "ccsb_conflict",
    "ccsb_publishsnapshot",
    "ccsb_scheduleitem",
    "ccsb_scheduleassignment",
    "ccsb_schedulerequirement",
    "ccsb_resource",
    "ccsb_service",
    "ccsb_location",
}

GAP_TABLES = [
    ("ccsb_publishlock", "No lock table/equivalent in dbmodel.xlsx; action plan marks this as a P0 V1 decision."),
    ("ccsb_publishsnapshotitem", "Snapshot header exists, but item-level managed before/after evidence is absent."),
    ("ccsb_schedulechange", "No pending/change ledger table is present."),
    ("ccsb_operationaudit or ccsb_auditlog", "Operation tables exist, but no append-only audit table is modeled."),
    ("ccsb_outboxevent", "No reserved outbox delivery table is present."),
    ("ccsb_scheduleactivity", "Activity is not modeled as a separate table; it may be intended as ccsb_scheduleitem.itemtype but needs approval."),
    ("ccsb_resourcecapability", "No normalized capability table is present; capability may be JSON-governed but needs approval."),
]


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    if not rows:
        return
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def parse_dbmodel() -> dict:
    wb = openpyxl.load_workbook(DBMODEL, read_only=True, data_only=True)
    entity_ws = wb["Entities list"]
    entities: dict[str, dict] = {}
    for row in entity_ws.iter_rows(min_row=4, values_only=True):
        if not row or not row[4]:
            continue
        logical = str(row[4]).lower()
        entities[logical] = {
            "display": row[0],
            "schema": row[3],
            "logical": logical,
            "object_type_code": row[5],
            "is_custom": bool(row[6]),
            "audit_enabled": bool(row[7]),
            "ownership": row[8],
            "description": row[2],
            "fields": {},
        }

    for ws in wb.worksheets:
        if ws.title == "Entities list":
            continue
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(c) if c is not None else "" for c in rows[0]]
        idx = {name: i for i, name in enumerate(headers) if name}
        field_names = {str(r[0]).lower() for r in rows[1:] if r and r[0]}
        logical = None
        for entity_name in entities:
            if f"{entity_name}id" in field_names:
                logical = entity_name
                break
        if logical is None:
            m = re.search(r"\(([^)]+)\)", ws.title)
            if m:
                logical = m.group(1).lower()
        if logical not in entities:
            continue
        for r in rows[1:]:
            if not r or not r[0]:
                continue
            field = str(r[0]).lower()
            entities[logical]["fields"][field] = {
                "schema": r[idx.get("Schema Name", 1)] if idx.get("Schema Name") is not None else "",
                "display": r[idx.get("Display Name", 2)] if idx.get("Display Name") is not None else "",
                "type": r[idx.get("Attribute Type", 3)] if idx.get("Attribute Type") is not None else "",
                "description": r[idx.get("Description", 4)] if idx.get("Description") is not None else "",
                "required": r[idx.get("Required Level", 7)] if idx.get("Required Level") is not None else "",
                "audit": r[idx.get("Audit Enabled", 9)] if idx.get("Audit Enabled") is not None else "",
                "secured": r[idx.get("Secured", 10)] if idx.get("Secured") is not None else "",
            }

    custom_entities = {name for name, e in entities.items() if name.startswith("ccsb_")}
    entity_audit_off = [name for name, e in entities.items() if name.startswith("ccsb_") and not e["audit_enabled"]]
    secured_custom_fields = [
        (entity, field)
        for entity, data in entities.items()
        if entity.startswith("ccsb_")
        for field, meta in data["fields"].items()
        if str(meta.get("secured", "")).lower() == "true"
    ]
    return {
        "entities": entities,
        "custom_entities": custom_entities,
        "entity_audit_off": entity_audit_off,
        "secured_custom_fields": secured_custom_fields,
        "missing_expected": sorted(EXPECTED_V1_TABLES - custom_entities),
    }


def classify_task(task: dict[str, str]) -> dict[str, str]:
    tid = task["TaskID"]
    if tid in COMPLETE_RULES:
        evidence, blocker = COMPLETE_RULES[tid]
        return {
            "RecommendedStatus": "Completed",
            "EvidenceLevel": "Direct dbmodel schema evidence",
            "Evidence": evidence,
            "RemainingCondition": blocker,
        }
    if tid in PROGRESS_RULES:
        evidence, blocker = PROGRESS_RULES[tid]
        return {
            "RecommendedStatus": "In Progress",
            "EvidenceLevel": "Partial dbmodel schema evidence",
            "Evidence": evidence,
            "RemainingCondition": blocker,
        }
    return {
        "RecommendedStatus": "Not Started",
        "EvidenceLevel": "No direct dbmodel evidence",
        "Evidence": "The current database model does not directly complete or materially start this task.",
        "RemainingCondition": "Requires runtime code, UX work, tests, ALM evidence, documentation, commercial work, or a missing schema decision outside the current dbmodel evidence.",
    }


def css() -> str:
    return """
:root{--ink:#17232d;--muted:#5d6d79;--canvas:#f5f7fa;--paper:#fff;--line:#d8e0e7;--nav:#102c3e;--blue:#176b98;--green:#237a52;--amber:#9b6512;--red:#af3d49;--bluebg:#eaf4fa;--greenbg:#ebf7f0;--amberbg:#fff6e6;--redbg:#fff0f2}
*{box-sizing:border-box}body{margin:0;background:var(--canvas);color:var(--ink);font:14px/1.55 "Segoe UI",Arial,sans-serif}
.layout{display:grid;grid-template-columns:292px minmax(0,1fr);min-height:100vh}.side{position:sticky;top:0;height:100vh;overflow:auto;background:var(--nav);color:#e9f4f8;padding:24px 18px}
.brand{padding-bottom:16px;margin-bottom:16px;border-bottom:1px solid #ffffff26}.brand b{display:block;font-size:18px}.brand span{display:block;color:#b7d0dd;font-size:12px;margin-top:5px}
.label{font-size:10px;letter-spacing:.12em;text-transform:uppercase;color:#91b7ca;margin:18px 8px 7px}.side a{display:block;color:#dcecf4;text-decoration:none;padding:8px 9px;border-radius:7px;font-size:13px}.side a:hover{background:#1d4b64}
.doc{max-width:1320px;margin:0 auto;padding:32px 42px 72px}.hero,.section{background:var(--paper);border:1px solid var(--line);border-radius:10px}.hero{border-left:8px solid var(--blue);padding:28px 30px;box-shadow:0 8px 26px #102c3e12}.section{padding:24px 26px;margin:18px 0}
h1{font-size:30px;line-height:1.16;margin:0 0 10px;color:#123c59}h2{font-size:22px;margin:0 0 12px;color:#123e5b}h3{font-size:16px;margin:18px 0 8px;color:#184e6d}p{margin:8px 0}.lead{font-size:16px;color:#354a5b;max-width:1040px}
.grid{display:grid;gap:12px}.four{grid-template-columns:repeat(4,minmax(0,1fr))}.three{grid-template-columns:repeat(3,minmax(0,1fr))}.two{grid-template-columns:repeat(2,minmax(0,1fr))}
.metric{border:1px solid var(--line);border-radius:9px;padding:14px;background:#fff}.num{font-size:28px;line-height:1;font-weight:800;color:var(--blue)}.metric .txt{font-size:11px;line-height:1.25;font-weight:800;text-transform:uppercase;letter-spacing:.05em;color:var(--muted);margin-top:6px}
.callout{border-left:5px solid var(--blue);background:var(--bluebg);padding:13px 15px;border-radius:0 8px 8px 0;margin:15px 0}.callout.red{border-color:var(--red);background:var(--redbg)}.callout.amber{border-color:var(--amber);background:var(--amberbg)}.callout.green{border-color:var(--green);background:var(--greenbg)}
.table{overflow:auto;border:1px solid var(--line);border-radius:9px;margin:12px 0}table{border-collapse:collapse;width:100%;min-width:880px;background:#fff}th,td{padding:8px 9px;text-align:left;vertical-align:top;border-right:1px solid var(--line);border-bottom:1px solid var(--line)}th{background:#edf5f9;color:#174b67;font-size:11px;text-transform:uppercase;letter-spacing:.03em}tr:last-child td{border-bottom:0}th:last-child,td:last-child{border-right:0}
.tag{display:inline-block;border-radius:999px;padding:3px 8px;font-weight:800;font-size:11px;white-space:nowrap}.complete{background:#dff2e7;color:#17613f}.progress{background:#fff0cb;color:#825000}.not{background:#eef2f6;color:#4f606d}.p0{background:#ffe0e4;color:#8a2b37}.p1{background:#fff0cb;color:#825000}.later{background:#dfedf7;color:#155d86}
code{font:12px Consolas,"Courier New",monospace;background:#eff4f7;border:1px solid #d4e0e8;border-radius:4px;padding:1px 4px}.small,.footer{font-size:12px;color:var(--muted)}
ul{margin:8px 0 8px 20px;padding:0}li{margin:5px 0}
@media(max-width:940px){.layout{display:block}.side{position:static;height:auto}.side nav{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:2px}.doc{padding:22px 16px}.four,.three,.two{grid-template-columns:1fr}}
"""


def status_tag(status: str) -> str:
    klass = {"Completed": "complete", "In Progress": "progress", "Not Started": "not"}.get(status, "not")
    return f'<span class="tag {klass}">{html.escape(status)}</span>'


def priority_tag(priority: str) -> str:
    klass = "later" if priority.lower() == "later" else priority.lower()
    return f'<span class="tag {klass}">{html.escape(priority)}</span>'


def td(value: object) -> str:
    return f"<td>{html.escape('' if value is None else str(value))}</td>"


def render_table(headers: list[str], rows: list[list[object]]) -> str:
    head = "".join(f"<th>{html.escape(h)}</th>" for h in headers)
    body = []
    for row in rows:
        body.append("<tr>" + "".join(td(v) for v in row) + "</tr>")
    return f'<div class="table"><table><tr>{head}</tr>{"".join(body)}</table></div>'


def render_html(schema: dict, task_rows: list[dict[str, str]], story_rows: list[dict[str, str]], story_summary: list[dict[str, str]]) -> str:
    task_counts = Counter(r["RecommendedStatus"] for r in task_rows)
    story_counts = Counter(r["RecommendedStatus"] for r in story_summary)
    changed_tasks = [r for r in task_rows if r["RecommendedStatus"] != r["CurrentStatus"]]
    complete_tasks = [r for r in task_rows if r["RecommendedStatus"] == "Completed"]
    progress_tasks = [r for r in task_rows if r["RecommendedStatus"] == "In Progress"]
    p0_progress = [r for r in progress_tasks if r["Priority"] == "P0"]

    story_table_rows = [
        [
            r["StoryID"],
            r["Priority"],
            r["Title"],
            r["RecommendedStatus"],
            r["CompletedTasks"],
            r["InProgressTasks"],
            r["NotStartedTasks"],
            r["Reason"],
        ]
        for r in story_summary
    ]

    changed_table_rows = [
        [
            r["TaskID"],
            r["ParentStoryID"],
            r["Priority"],
            r["TaskCategory"],
            r["Title"],
            r["RecommendedStatus"],
            r["Evidence"],
            r["RemainingCondition"],
        ]
        for r in changed_tasks
    ]

    completed_rows = [
        [r["TaskID"], r["ParentStoryID"], r["Title"], r["Evidence"], r["RemainingCondition"]]
        for r in complete_tasks
    ]
    progress_rows = [
        [r["TaskID"], r["ParentStoryID"], r["Priority"], r["Title"], r["Evidence"], r["RemainingCondition"]]
        for r in progress_tasks
    ]
    p0_next_rows = [
        [r["TaskID"], r["ParentStoryID"], r["Title"], r["RemainingCondition"]]
        for r in p0_progress[:35]
    ]
    gaps_rows = GAP_TABLES

    all_task_rows = [
        [
            r["TaskID"],
            r["ParentStoryID"],
            r["Priority"],
            r["TaskCategory"],
            r["Title"],
            r["CurrentStatus"],
            r["RecommendedStatus"],
            r["EvidenceLevel"],
            r["RemainingCondition"],
        ]
        for r in task_rows
    ]

    nav = """
<div class="label">Summary</div>
<a href="#decision">Decision</a>
<a href="#counts">Counts</a>
<a href="#method">Method</a>
<div class="label">Evidence</div>
<a href="#schema">Schema Evidence</a>
<a href="#gaps">Schema Gaps</a>
<div class="label">Backlog</div>
<a href="#stories">Story Status</a>
<a href="#completed">Completed Tasks</a>
<a href="#inprogress">In-Progress Tasks</a>
<a href="#alltasks">All Tasks</a>
<a href="#plan">Continuation Plan</a>
"""

    return f"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<title>CCSB dbmodel Backlog Status Overlay</title>
<style>{css()}</style>
</head>
<body>
<div class="layout">
<aside class="side"><div class="brand"><b>CCSB dbmodel Status Overlay</b><span>Backlog comparison from current database model<br>Prepared 2026-07-03</span></div><nav>{nav}</nav></aside>
<main class="doc">
<section id="decision" class="hero">
<div class="small" style="text-transform:uppercase;letter-spacing:.1em;font-weight:800;color:#176b98">Corrected interpretation</div>
<h1>Current Database Model vs. User Stories and Tasks</h1>
<p class="lead">This report compares <code>metadata/dbmodel.xlsx</code> to the task backlog and recommends a status overlay. It does not simply repeat the CSV status fields. It also respects the post-validation action plan: <strong>do not mark US-004 complete wholesale</strong> until schema hardening, audit/security, migration, UX pilot, and clean import/export evidence exist.</p>
<div class="callout amber"><strong>Recommendation:</strong> mark database-model-complete tasks as complete, move schema-supported but unfinished tasks to in progress, and leave runtime/API/UX/test/release tasks as not started unless there is direct evidence outside the database model.</div>
</section>

<section id="counts" class="section">
<h2>1. Recommended Status Counts</h2>
<div class="grid four">
<div class="metric"><div class="num">{len(task_rows)}</div><div class="txt">Tasks reviewed</div></div>
<div class="metric"><div class="num">{task_counts['Completed']}</div><div class="txt">Tasks can be completed</div></div>
<div class="metric"><div class="num">{task_counts['In Progress']}</div><div class="txt">Tasks can be in progress</div></div>
<div class="metric"><div class="num">{story_counts['In Progress']}</div><div class="txt">Stories can be in progress</div></div>
</div>
<div class="grid three" style="margin-top:12px">
<div class="metric"><div class="num">{story_counts['Completed']}</div><div class="txt">Stories can be completed</div></div>
<div class="metric"><div class="num">{story_counts['Not Started']}</div><div class="txt">Stories remain not started</div></div>
<div class="metric"><div class="num">{len(changed_tasks)}</div><div class="txt">Task status changes proposed</div></div>
</div>
<div class="callout red"><strong>Important:</strong> no user story should be marked complete from dbmodel evidence alone. Each story still has unresolved implementation, validation, UX, testing, security, deployment, or approval tasks.</div>
</section>

<section id="method" class="section">
<h2>2. Method</h2>
<ul>
<li><strong>Completed</strong> means the current database model directly satisfies the schema/data-model deliverable for that task, with no known dbmodel gap in the task's core data shape.</li>
<li><strong>In Progress</strong> means the dbmodel contains meaningful foundation evidence, but the task still needs missing fields/tables, approval, server logic, UX, tests, security, migration or live Dataverse proof.</li>
<li><strong>Not Started</strong> means the database model does not directly prove progress on that task.</li>
</ul>
<p class="small">The source backlog files were not edited. Use the generated CSV overlays if you want to import the proposed status changes after review.</p>
</section>

<section id="schema" class="section">
<h2>3. Schema Evidence Read From dbmodel.xlsx</h2>
<div class="grid four">
<div class="metric"><div class="num">{len(schema['custom_entities'])}</div><div class="txt">Custom ccsb tables</div></div>
<div class="metric"><div class="num">{sum(len(e['fields']) for e in schema['entities'].values())}</div><div class="txt">Workbook fields</div></div>
<div class="metric"><div class="num">{len(schema['entity_audit_off'])}</div><div class="txt">ccsb tables with entity audit off</div></div>
<div class="metric"><div class="num">{len(schema['secured_custom_fields'])}</div><div class="txt">Secured custom fields</div></div>
</div>
<div class="callout amber">The workbook has many field-level audit flags set, but every custom table's entity-level audit flag is disabled. There are no custom <code>ccsb_*</code> secured fields in the model, so security/audit tasks cannot be closed.</div>
</section>

<section id="gaps" class="section">
<h2>4. Schema Gaps Blocking Broader Completion</h2>
{render_table(["Gap", "Reason"], gaps_rows)}
</section>

<section id="stories" class="section">
<h2>5. User Story Status Overlay</h2>
{render_table(["Story", "Priority", "Title", "Recommended status", "Completed tasks", "In-progress tasks", "Not-started tasks", "Reason"], story_table_rows)}
</section>

<section id="completed" class="section">
<h2>6. Tasks That Can Be Marked Completed</h2>
{render_table(["Task", "Story", "Title", "Evidence", "Remaining note"], completed_rows)}
</section>

<section id="inprogress" class="section">
<h2>7. Tasks That Can Be Marked In Progress</h2>
{render_table(["Task", "Story", "Priority", "Title", "Evidence", "Remaining condition"], progress_rows)}
</section>

<section id="alltasks" class="section">
<h2>8. All Task Status Recommendations</h2>
{render_table(["Task", "Story", "Priority", "Category", "Title", "Current status", "Recommended status", "Evidence level", "Remaining condition"], all_task_rows)}
</section>

<section id="plan" class="section">
<h2>9. Prioritized Continuation Plan</h2>
<div class="callout red"><strong>P0 first:</strong> close the missing V1 control model before scaling UX or runtime work: publish lock, snapshot item, schedule change ledger, audit/correlation model, outbox reservation, activity/capability decision, field security and entity audit enablement.</div>
{render_table(["Task", "Story", "Title", "Next condition to clear"], p0_next_rows)}
<p class="footer">Generated from <code>{html.escape(str(DBMODEL.relative_to(ROOT)))}</code>, backlog CSVs under <code>Black_Logs</code>, and the post-validation action plan at <code>{html.escape(str(ACTION_PLAN))}</code>.</p>
</section>
</main></div></body></html>
"""


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    schema = parse_dbmodel()
    tasks = read_csv(TASKS_CSV)
    stories = read_csv(STORIES_CSV)

    task_rows: list[dict[str, str]] = []
    for t in tasks:
        cls = classify_task(t)
        task_rows.append(
            {
                "TaskID": t["TaskID"],
                "ParentStoryID": t["ParentStoryID"],
                "EpicID": t["EpicID"],
                "Priority": t["Priority"],
                "Phase": t["Phase"],
                "TaskCategory": t["TaskCategory"],
                "Title": t["Title"],
                "CurrentStatus": t["Status"],
                **cls,
            }
        )

    tasks_by_story: dict[str, list[dict[str, str]]] = defaultdict(list)
    for r in task_rows:
        tasks_by_story[r["ParentStoryID"]].append(r)

    story_summary: list[dict[str, str]] = []
    for s in stories:
        rows = tasks_by_story.get(s["StoryID"], [])
        counts = Counter(r["RecommendedStatus"] for r in rows)
        if rows and counts["Completed"] == len(rows):
            rec = "Completed"
            reason = "All child tasks are complete by dbmodel evidence."
        elif counts["Completed"] or counts["In Progress"]:
            rec = "In Progress"
            reason = f"dbmodel evidence supports {counts['Completed']} completed and {counts['In Progress']} in-progress child tasks, but remaining tasks are not complete."
        else:
            rec = "Not Started"
            reason = "No direct dbmodel evidence moves this story."
        story_summary.append(
            {
                "StoryID": s["StoryID"],
                "EpicID": s["EpicID"],
                "Priority": s["Priority"],
                "Title": s["Title"],
                "CurrentStatus": s["Status"],
                "RecommendedStatus": rec,
                "CompletedTasks": str(counts["Completed"]),
                "InProgressTasks": str(counts["In Progress"]),
                "NotStartedTasks": str(counts["Not Started"]),
                "Reason": reason,
            }
        )

    html_text = render_html(schema, task_rows, stories, story_summary)
    html_path = OUT_DIR / "ccsb-dbmodel-backlog-status-overlay.html"
    html_path.write_text(html_text, encoding="utf-8")

    write_csv(OUT_DIR / "ccsb-dbmodel-task-status-overlay.csv", task_rows)
    write_csv(OUT_DIR / "ccsb-dbmodel-story-status-overlay.csv", story_summary)

    summary = {
        "tasks": Counter(r["RecommendedStatus"] for r in task_rows),
        "stories": Counter(r["RecommendedStatus"] for r in story_summary),
        "files": {
            "html": str(html_path),
            "task_csv": str(OUT_DIR / "ccsb-dbmodel-task-status-overlay.csv"),
            "story_csv": str(OUT_DIR / "ccsb-dbmodel-story-status-overlay.csv"),
        },
    }
    print(json.dumps(summary, indent=2, default=dict))


if __name__ == "__main__":
    main()
